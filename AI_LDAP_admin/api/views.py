from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from ldap3 import *
import json, re, random 
from django.contrib.auth.models import User, Group
from .models import UserDetail

import base64
from passlib.hash import ldap_md5_crypt

from rest_framework.parsers import JSONParser
from rest_framework.response import Response
from rest_framework.decorators import api_view
from .serializers import UserSerializer, GroupSerializer

salt = 'cguadmin'

def get_gid():
    while True:
        uid = random.randint(10000, 65535)  # Generate a random UID within the range of user IDs
'''        try:
            pwd.getgrgid(uid)  # Attempt to get the user entry for the generated UID
        except KeyError:
            return uid
'''
# connect to LDAP server
def connectLDAP():
    server = Server('ldap://120.126.23.245:31979')
    conn = Connection(server, user='cn=admin,dc=example,dc=org', password='Not@SecurePassw0rd', auto_bind=True)
    return conn

@api_view(['POST'])
def lab_list(request):
    data = json.loads(request.body.decode('utf-8'))
    user = data['user']
    group_list = []
    if user == 'root':
        conn = connectLDAP()
        conn.search('dc=example,dc=org', '(objectclass=posixGroup)', attributes=['cn'])
        for entry in conn.entries:
            group_list.append(entry.cn.value)
        conn.unbind()
    else:
        # check the user permission is one of the lab admin, return the lab name, in which the user is admin
        conn = connectLDAP()
        user_dn = 'cn={},ou=users,dc=example,dc=org'.format(user)
        conn.search(user_dn, '(objectclass=posixAccount)', attributes=['Description'])
        for entry in conn.entries:
            permission_list = entry.Description.values
            for permission in permission_list:
                if re.match(r'.*admin', str(permission)):
                    group_list.append(permission[:-5])
        conn.unbind()
        
    return Response(group_list, status=200)

@api_view(['GET'])
def user_list(request):
    conn = connectLDAP()
    # objectclass is posixAccount and top 
    conn.search('dc=example,dc=org', '(objectclass=posixAccount)', attributes=['cn'])
    user_list = []
    for entry in conn.entries:
        user_list.append(entry.cn.value)
    conn.unbind()
    return Response(user_list, status=200)


@api_view(['POST'])
def get_group_corresponding_user(request):
    # get all group and corresponding user
    data = json.loads(request.body.decode('utf-8'))
    user = data['user']
    conn = connectLDAP()
    user_dn = 'cn={},ou=users,dc=example,dc=org'.format(user)
    conn.search(user_dn, '(objectclass=posixAccount)', attributes=['Description'])
    group_list = []
    permission_list = []
    for entry in conn.entries:
        permission_list = entry.Description.values
    for permission in permission_list:
        if (permission == 'admin'):
            # get all attributes
            conn.search('dc=example,dc=org', '(objectclass=posixGroup)', attributes=['*'])
            for entry in conn.entries:
                member_uids = []
                group_dn = entry.cn.value
                try:
                    member_uids_entry = entry.memberUid.values
                    for member_uid in member_uids_entry:
                        member_uids.append(str(member_uid))

                    # Append the group cn and corresponding memberUids
                    group_list.append({
                        'group_dn': group_dn,
                        'member_uids': member_uids
                    })
                except:
                    group_list.append({
                        'group_dn': group_dn,
                        'member_uids': []
                })
        # permission is specificed lab admin with name ex (Leeadmin)
        elif re.match(r'.*admin', str(permission)):
            labname = permission[:-5]
            conn.search('cn={},ou=Groups,dc=example,dc=org'.format(labname), '(objectclass=posixGroup)', attributes=['*'])
            for entry in conn.entries:
                member_uids = []
                group_dn = entry.cn.value
                try:
                    member_uids_entry = entry.memberUid.values
                    for member_uid in member_uids_entry:
                        member_uids.append(str(member_uid))

                    # Append the group cn and corresponding memberUids
                    group_list.append({
                        'group_dn': group_dn,
                        'member_uids': member_uids
                    })
                except:
                    group_list.append({
                        'group_dn': group_dn,
                        'member_uids': []
                })
        # permission is specificed lab user with name ex (Lee)
        else:
            conn.search('cn={},ou=Groups,dc=example,dc=org'.format(permission), '(objectclass=posixGroup)', attributes=['*'])
            for entry in conn.entries:
                # get the group name and only one memberUid of this user
                group_dn = entry.cn.value
                try:
                    member_uids = [user]
                    # Append the group cn and corresponding memberUids
                    group_list.append({
                        'group_dn': group_dn,
                        'member_uids': member_uids
                    })
                except:
                    group_list.append({
                        'group_dn': group_dn,
                        'member_uids': []
                })
    
    if (data['user'] == 'root'):
        conn.search('dc=example,dc=org', '(objectclass=posixGroup)', attributes=['*'])
        for entry in conn.entries:
            member_uids = []
            group_dn = entry.cn.value
            try:
                member_uids_entry = entry.memberUid.values
                for member_uid in member_uids_entry:
                    member_uids.append(str(member_uid))

                # Append the group cn and corresponding memberUids
                group_list.append({
                    'group_dn': group_dn,
                    'member_uids': member_uids
                })
            except:
                group_list.append({
                    'group_dn': group_dn,
                    'member_uids': []
            })

    conn.unbind()
    return Response(group_list, status=200)

@api_view(['POST'])
def get_lab_info(request):
    data = json.loads(request.body.decode('utf-8'))
    labname = data['lab']
    conn = connectLDAP()
    conn.search('cn={},ou=Groups,dc=example,dc=org'.format(labname), '(objectclass=posixGroup)', attributes=['*'])
    data = {}
    for entry in conn.entries:
        try:
            data = {
                "cn": entry.cn.value,
                "gidNumber": entry.gidNumber.value,
                "memberUid": entry.memberUid.values
            }
        except:
            data = {
                "cn": entry.cn.value,
                "gidNumber": entry.gidNumber.value,
                "memberUid": []
            }

    conn.unbind()
    return Response(data, status=200)

@api_view(['POST'])
def addlab(request):
    data = json.loads(request.body.decode('utf-8'))
    labname = data['lab']
    group_dn = 'cn={},ou=Groups,dc=example,dc=org'.format(labname)
    conn = connectLDAP()
    conn.add('cn={},ou=Groups,dc=example,dc=org'.format(labname), ['posixGroup', 'top'], {'cn': ['{}'.format(labname)], 'gidNumber': ['1001']})
    group = Group.objects.create(name=labname)
    # add all permission to the group
    group.save()
    
    conn.unbind()
    return Response(status=200)

@api_view(['POST'])
def adduser(request):
    data = json.loads(request.body.decode('utf-8'))
    username = data['username']
    firstname = data['first_name']
    lastname = data['last_name']
    password = data['password']
    labname = data['lab']
    email = data['email']
    group_dn = 'cn={},ou=Groups,dc=example,dc=org'.format(labname)
    user_dn = 'cn={},ou=users,dc=example,dc=org'.format(username),
    conn = connectLDAP()
    conn.add(user_dn, ['inetOrgPerson', 'posixAccount', 'shadowAccount', 'top'],
        {'cn': username, 'givenName': username, 'sn' : username ,
        'uid': username, 'uidNumber': '2001', 'gidNumber': '1001', "mail": email,
        'homeDirectory': '/home/{}'.format(username), 'loginShell': '/bin/bash',
        'userPassword': ldap_md5_crypt.hash(password, salt=salt), 'shadowFlag': '0', 'shadowMin': '0', 'shadowMax': '99999', 
        'shadowWarning': '0', 'shadowInactive': '99999', 'shadowLastChange': '12011', 
        'shadowExpire': '99999', 'Description': [labname]})
    user = User.objects.create_user(username=username, password=password, first_name=firstname, last_name=lastname, email=data['email'])
    if data['lab'] is not None:
        group_dn = 'cn={},ou=Groups,dc=example,dc=org'.format(labname)
        conn.modify(group_dn, {'memberUid': [(MODIFY_ADD, [username])]})
        user.groups.add(Group.objects.get(name=labname))
        if data['is_lab_manager'] is False:
            user.userdetail_set.create(uid=user, labname=Group.objects.get(name=labname), permission=2)
            conn.modify(user_dn, {'Description': [(MODIFY_ADD, [labname])]})
        elif data['is_lab_manager'] is True:
            user.userdetail_set.create(uid=user, labname=Group.objects.get(name=labname), permission=1)
            conn.modify(user_dn, {'Description': [(MODIFY_ADD, ['{}admin'.format(labname)])]})
            conn.modify(user_dn, {'Description': [(MODIFY_DELETE, [labname])]})
    conn.unbind()
    user.save()
    
    return Response(status=200)

@api_view(['POST'])
def add_admin(request):
    data = json.loads(request.body.decode('utf-8'))
    username = data['username']
    user_dn = 'cn={},ou=users,dc=example,dc=org'.format(username),
    conn = connectLDAP()
    conn.modify(user_dn, {'Description': [(MODIFY_ADD, ['admin'])]})
    user = User.objects.get(username=username)
    detail = UserDetail.objects.get(uid=user)
    detail.permission = 0
    try:
        detail.labname = Group.objects.get(name='root')
        detail.save()
    except:
        pass
    # make user to be superuser
    user.is_superuser = True
    user.is_staff = True
    user.save()
    
    # ldap admin 
    return Response(status=200)

def syschronize_ldap(requset):
    conn = connectLDAP()
    conn.search('dc=example,dc=org', '(objectclass=posixGroup)', attributes=['cn'])
    group_list = []
    for entry in conn.entries:
        group_list.append(entry.entry_dn)
    conn.search('dc=example,dc=org', '(objectclass=posixAccount)', attributes=['cn'])
    account_list = []
    for entry in conn.entries:
        account_list.append(entry.entry_gidNumber)
        conn.unbind()
    # get the user with corresponding group
    
    return JsonResponse({'group_list': group_list, 'account_list': account_list}, status=200)

@api_view(['POST'])
def get_user_info(request):
    data = json.loads(request.body.decode('utf-8'))
    conn = connectLDAP()
    conn.search('cn={},ou=users,dc=example,dc=org'.format(data['username']), '(objectclass=posixAccount)', attributes=['*'])
    user = User.objects.get(username=data['username'])
    data = {
        "username": conn.entries[0].cn.value,
        "first_name": conn.entries[0].givenName.value,
        "last_name": conn.entries[0].sn.value,
        "email": conn.entries[0].mail.value,
    }
    return Response(data, status=200)    

@api_view(['POST'])
def user_delete(request):
    data = json.loads(request.body.decode('utf-8'))
    username = data['username']
    conn = connectLDAP()
    conn.delete('cn={},ou=users,dc=example,dc=org'.format(username))
    ## delete the user memberUID from the group
    conn.search('dc=example,dc=org', '(objectclass=posixGroup)', attributes=['cn'])
    for entry in conn.entries:
        try:
            conn.modify(entry.entry_dn, {'memberUid': [(MODIFY_DELETE, [username])]})
        except:
            pass
    User.objects.get(username=username).delete()
    return Response(status=200)

@api_view(['POST'])
def lab_delete(request):
    data = json.loads(request.body.decode('utf-8'))
    labname = data['lab']
    user_list = User.objects.filter(groups__name=labname)
    for user in user_list:
        if(len(user.groups.all()) == 1):
            user.delete()
        else:
            user.groups.remove(Group.objects.get(name=labname))
    #Group.objects.get(name=labname)
    conn = connectLDAP()
    conn.delete('cn={},ou=Groups,dc=example,dc=org'.format(labname))
    conn.unbind()
    return Response(status=200)
    
def user_group_num(requset):
    conn = connectLDAP()
    group_list = []
    user_list = []
    conn.search('dc=example,dc=org', '(objectclass=posixGroup)', attributes=['cn'])
    for entry in conn.entries:
        group_list.append(entry.cn.value)
    conn.search('dc=example,dc=org', '(objectclass=posixAccount)', attributes=['cn'])
    for entry in conn.entries:
        user_list.append(entry.cn.value)
    conn.unbind()
    # return the number of group and user
    data = {'lab_num': len(group_list), 'user_num': len(user_list)}
    return JsonResponse(data, safe=False)


@api_view(['POST'])
def add_lab_admin(request):
    data = json.loads(request.body.decode('utf-8'))
    username = data['username']
    labname = data['lab']
    User.objects.get(username=username).is_staff = True
    detail = UserDetail.objects.filter(uid=User.objects.get(username=username), labname=Group.objects.get(name=labname))
    detail.permission = 1
    detail.save()
    conn = connectLDAP()
    lab = conn.search('cn={},ou=Groups,dc=example,dc=org'.format(labname), '(objectclass=posixGroup)', attributes=['cn'])
    user = conn.search('cn={},ou=users,dc=example,dc=org'.format(username), '(objectclass=posixAccount)', attributes=['*'])
    for entry in user.entries:
        conn.modify(entry.entry_dn, {'Description': [(MODIFY_ADD, ['{}admin'.format(labname)])]})
    conn.unbind()

@api_view(['GET'])
def synchronize(request):
    # use to check if ldap and django data same
    conn = connectLDAP()

@api_view(['POST'])
def change_password(request):
    data = json.loads(request.body.decode('utf-8'))
    username = data['username']
    password = data['password']
    conn = connectLDAP()
    conn.search('cn={},ou=users,dc=example,dc=org'.format(username), '(objectclass=posixAccount)', attributes=['*'])
    try:
        for entry in conn.entries:
            conn.modify(entry.entry_dn, {'userPassword':[(MODIFY_REPLACE, [ldap_md5_crypt.hash(password, salt=salt)])]})
        conn.unbind
        user = User.objects.get(username=username)
        user.set_password(password)
        user.save()
        return Response(status=200)
    except:
        return Response(status=500)

@api_view(['POST'])
def change_user_info(request):
    data = json.loads(request.body.decode('utf-8'))
    username = data['username']
    firstname = data['firstname']
    lastname = data['lastname']
    email = data['email']
    try:
        conn = connectLDAP()
        conn.search('cn={},ou=users,dc=example,dc=org'.format(username), '(objectclass=posixAccount)', attributes=['*'])
        for entry in conn.entries:
            conn.modify(entry.entry_dn, {'givenName':[(MODIFY_REPLACE, [firstname])]})
            conn.modify(entry.entry_dn, {'sn':[(MODIFY_REPLACE, [lastname])]})
            conn.modify(entry.entry_dn, {'mail': [(MODIFY_REPLACE, [email])]})
        user = User.objects.get(username=username)
        user.first_name = firstname
        user.last_name = lastname
        user.email = email
        user.save()
        return Response(status=200)
    except:
        return Response(status=500)

from openpyxl import Workbook
from openpyxl import load_workbook

import xlrd
@api_view(['POST'])
def excel(request):
    # get the excel file from frontend
    excel_file = request.FILES['file']
    print(excel_file)
    wb = load_workbook(excel_file)
    sheet = wb.active
    # get the data from excel attribute["Username","Group","password","email", "firstname", "lastname", "permission"]
    user_list = []
    for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        user_list.append(row)
    # connect to ldap
    conn = connectLDAP()
    # add user to ldap
    for user in user_list:
        # if user is not exist and group is not exist, add new user and group
        if not conn.search('cn={},ou=users,dc=example,dc=org'.format(user[0]), '(objectclass=posixAccount)', attributes=['*']) and not conn.search('cn={},ou=Groups,dc=example,dc=org'.format(user[1]), '(objectclass=posixGroup)', attributes=['*']):
            conn.add('cn={},ou=users,dc=example,dc=org'.format(user[0]), ['inetOrgPerson', 'posixAccount', 'shadowAccount', 'top'],
                {'cn': user[0], 'givenName': user[4], 'sn' : user[5] ,
                'uid': user[0], 'uidNumber': '2001', 'gidNumber': '1001', "mail": user[3],
                'homeDirectory': '/home/{}'.format(user[0]), 'loginShell': '/bin/bash',
                'userPassword': ldap_md5_crypt.hash(user[2], salt=salt), 'shadowFlag': '0', 'shadowMin': '0', 'shadowMax': '99999', 
                'shadowWarning': '0', 'shadowInactive': '99999', 'shadowLastChange': '12011', 
                'shadowExpire': '99999', 'Description': [user[1]]})
            conn.add('cn={},ou=Groups,dc=example,dc=org'.format(user[1]), ['posixGroup', 'top'], {'cn': ['{}'.format(user[1])], 'gidNumber': ['1001']})
            user = User.objects.create_user(username=user[0], password=user[2], first_name=user[4], last_name=user[5], email=user[3])
            user.groups.add(Group.objects.get(name=user[1]))
            user.userdetail_set.create(uid=user, labname=Group.objects.get(name=user[1]), permission=2)
            conn.modify('cn={},ou=Groups,dc=example,dc=org'.format(user[1]), {'memberUid': [(MODIFY_ADD, [user[0]])]})
            conn.unbind()
            user.save()
        # if user is not exist and group is exist, add new user and add user to group
        elif not conn.search('cn={},ou=users,dc=example,dc=org'.format(user[0]), '(objectclass=posixAccount)', attributes=['*']) and conn.search('cn={},ou=Groups,dc=example,dc=org'.format(user[1]), '(objectclass=posixGroup)', attributes=['*']):
            conn.add('cn={},ou=users,dc=example,dc=org'.format(user[0]), ['inetOrgPerson', 'posixAccount', 'shadowAccount', 'top'],
                {'cn': user[0], 'givenName': user[4], 'sn' : user[5] ,
                'uid': user[0], 'uidNumber': '2001', 'gidNumber': '1001', "mail": user[3],
                'homeDirectory': '/home/{}'.format(user[0]), 'loginShell': '/bin/bash',
                'userPassword': ldap_md5_crypt.hash(user[2], salt=salt), 'shadowFlag': '0', 'shadowMin': '0', 'shadowMax': '99999', 
                'shadowWarning': '0', 'shadowInactive': '99999', 'shadowLastChange': '12011', 
                'shadowExpire': '99999', 'Description': [user[1]]})
            user = User.objects.create_user(username=user[0], password=user[2], first_name=user[4], last_name=user[5], email=user[3])
            user.groups.add(Group.objects.get(name=user[1]))
            user.userdetail_set.create(uid=user, labname=Group.objects.get(name=user[1]), permission=2)
            conn.modify('cn={},ou=Groups,dc=example,dc=org'.format(user[1]), {'memberUid': [(MODIFY_ADD, [user[0]])]})
            conn.unbind()
            user.save()
        # if user is exist and group is not exist, add new group and add user to group
        elif conn.search('cn={},ou=users,dc=example,dc=org'.format(user[0]), '(objectclass=posixAccount)', attributes=['*']) and not conn.search('cn={},ou=Groups,dc=example,dc=org'.format(user[1]), '(objectclass=posixGroup)', attributes=['*']):
            conn.add('cn={},ou=Groups,dc=example,dc=org'.format(user[1]), ['posixGroup', 'top'], {'cn': ['{}'.format(user[1])], 'gidNumber': ['1001']})
            user = User.objects.create_user(username=user[0], password=user[2], first_name=user[4], last_name=user[5], email=user[3])
            user.groups.add(Group.objects.get(name=user[1]))
            user.userdetail_set.create(uid=user, labname=Group.objects.get(name=user[1]), permission=2)
            conn.modify('cn={},ou=Groups,dc=example,dc=org'.format(user[1]), {'memberUid': [(MODIFY_ADD, [user[0]])]})
            conn.unbind()
            user.save()
        # if user is exist and group is exist, add user to group
        elif conn.search('cn={},ou=users,dc=example,dc=org'.format(user[0]), '(objectclass=posixAccount)', attributes=['*']) and conn.search('cn={},ou=Groups,dc=example,dc=org'.format(user[1]), '(objectclass=posixGroup)', attributes=['*']):
            user = User.objects.create_user(username=user[0], password=user[2], first_name=user[4], last_name=user[5], email=user[3])
            user.groups.add(Group.objects.get(name=user[1]))
            user.userdetail_set.create(uid=user, labname=Group.objects.get(name=user[1]), permission=2)
            conn.modify('cn={},ou=Groups,dc=example,dc=org'.format(user[1]), {'memberUid': [(MODIFY_ADD, [user[0]])]})
            conn.unbind()
            user.save()

    return Response(status=200)
        


def get_permission(user, group):
    conn = connectLDAP()
    conn.search('cn={},ou=users,dc=example,dc=org'.format(user), '(objectclass=posixAccount)', attributes=['Description'])
    for entry in conn.entries:
        if re.match(r'{}admin'.format(group), str(entry.Description.value)):
            return "admin"
    return "user"

@api_view(['GET'])
def export_ldap(request):
    conn = connectLDAP()
    conn.search('dc=example,dc=org', '(objectclass=posixGroup)', attributes=['*'])
    user_list = []
    user_list.append(["Username","Group","password","email", "firstname", "lastname", "permission"])
    for entry in conn.entries:
        ## if group without memberUid attribute, skip
        try:
            group = entry.cn.value
            member_list = entry.memberUid.value
            
            # check group has muti user or not
            if isinstance(member_list, list):
                for member_entry in member_list:
                    conn.search('cn={},ou=users,dc=example,dc=org'.format(member_entry), '(objectclass=posixAccount)', attributes=['*'])
                    for user_entry in conn.entries:
                        user_list.append([user_entry.cn.value, group, user_entry.userPassword.value, user_entry.mail.value, user_entry.givenName.value, user_entry.sn.value, get_permission(user_entry.cn.value, group)])
            elif isinstance(member_list, str):
                conn.search('cn={},ou=users,dc=example,dc=org'.format(member_list), '(objectclass=posixAccount)', attributes=['*'])
                for user_entry in conn.entries:
                    user_list.append([user_entry.cn.value, group, user_entry.userPassword.value, user_entry.mail.value, user_entry.givenName.value, user_entry.sn.value, get_permission(user_entry.cn.value, group)])
        except:
            continue
        
    # make data to excel
    workbook = Workbook()
    worksheet = workbook.active
    for row in user_list:
        worksheet.append(row)
    workbook.save("data.xlsx")
    excel_file_path = 'data.xlsx'
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    workbook.save(response)
    
    return response


    